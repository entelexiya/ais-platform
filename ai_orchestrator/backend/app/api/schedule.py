from fastapi import APIRouter, Response, Body, Depends, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import pandas as pd
import os
import tempfile
from sqlalchemy.orm import Session
from app.db.database import get_db
from app.db.models import (
    Teacher, Room, Parallel, Class, LentaGroup,
    TimeSlot, ScheduleEntry, StaffScheduleEntry
)
from app.services.scheduler import find_substitution, process_teacher_absence_event
from app.services.legal_generator import generate_substitution_order
from app.services.pdf_service import create_order_html
from app.data_loader import load_staff, load_schedule
from app.services.generator import generate_weekly_schedule
from app.services.conflict_checker import ConflictChecker
from app.services.staff_scheduler import (
    get_staff_weekly_schedule, inject_task_into_schedule, get_all_staff_list
)

router = APIRouter()

class ScheduleTarget(BaseModel):
    classes: Optional[List[str]] = None

import re
from openpyxl.styles import Alignment, PatternFill, Font

class ScheduleExportData(BaseModel):
    schedule: List[dict]

@router.post("/download-excel")
async def export_excel(data: ScheduleExportData):
    """
    Принимает JSON сгенерённого расписания и отдаёт "Ленточное расписание"
    Сгруппировано по Параллелям и Дням недели, в формате Pivot-матрицы.
    """
    df = pd.DataFrame(data.schedule)
    tmp_path = os.path.join(tempfile.gettempdir(), f"ribbon_schedule_{int(datetime.now().timestamp())}.xlsx")
    
    if df.empty:
        df.to_excel(tmp_path, index=False, engine='openpyxl')
    else:
        # 1. Извлекаем параллель
        def get_grade(x):
            match = re.search(r'\d+', str(x))
            return int(match.group()) if match else 0
            
        df['Grade'] = df['Класс'].apply(get_grade)
        
        # 2. Формируем содержимое ячейки с переносом строк
        df['Cell'] = df['Предмет'] + "\n" + df['Учитель'] + "\n(" + df['Кабинет'] + ")"
        
        with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
            grades_list = sorted(df['Grade'].unique())
            days_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
            
            for grade in grades_list:
                for day in days_order:
                    subset = df[(df['Grade'] == grade) & (df['День'] == day)]
                    if subset.empty:
                        continue
                        
                    # Разворачиваем Ленту (Pivot)
                    pivot = subset.pivot(index='Урок', columns='Класс', values='Cell')
                    
                    # Фиксируем порядок строк (уроки с 1 до 6)
                    pivot = pivot.reindex(list(range(1, 7)))
                    
                    # Имя вкладки (макс 31 символ в Excel) e.g., "10кл Пн"
                    day_short = day[:2]
                    sheet_name = f"{grade} кл. {day_short}"
                    pivot.to_excel(writer, sheet_name=sheet_name)
                    
                    # --- КРАСИВОЕ ФОРМАТИРОВАНИЕ OpenPyXL ---
                    worksheet = writer.sheets[sheet_name]
                    
                    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Tailwind Blue 500
                    header_font = Font(color="FFFFFF", bold=True)
                    index_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # Tailwind Blue 50
                    
                    for col in worksheet.columns:
                        col_letter = col[0].column_letter
                        worksheet.column_dimensions[col_letter].width = 26
                        
                        for cell in col:
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                            
                            # Форматируем шапку (Классы)
                            if cell.row == 1:
                                cell.fill = header_fill
                                cell.font = header_font
                            
                            # Форматируем боковик (Уроки)
                            if cell.column == 1 and cell.row > 1:
                                cell.fill = index_fill
                                cell.font = Font(bold=True, color="1E3A8A")
                                
                    # Высота строк для переноса
                    for row in range(2, 9):
                        worksheet.row_dimensions[row].height = 65

    return FileResponse(
        tmp_path, 
        filename="Aqbobek_Ленточное_Расписание.xlsx", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/generate-schedule")
async def api_generate_schedule(target: ScheduleTarget, db: Session = Depends(get_db)):
    """
    Генерирует расписание на неделю с нуля.
    Классы передаются в body: {"classes": ["1А", "5Б"]}.
    Результат возвращается в виде JSON-списка.
    """
    result = generate_weekly_schedule(target.classes, db, include_summary=True)
    schedule = result["schedule"]
    return {
        "status": "success",
        "total_slots": len(schedule),
        "summary": result["summary"],
        "schedule": schedule,
    }

@router.get("/teacher-profile")
async def get_teacher_profile(teacher_name: str = "Иванова И. И."):
    """Mock-профиль учителя, берет реальные данные из Excel"""
    try:
        staff = load_staff()
        schedule = load_schedule()

        # Find teacher
        teacher_info = next((t for t in staff if teacher_name.lower() in t.get("ФИО", "").lower()), None)
        if not teacher_info and staff:
            # Fallback to first teacher if not found
            teacher_info = staff[0]
            teacher_name = teacher_info.get("ФИО", "Неизвестно")

        # Find schedule for today (Monday mock)
        today_val = "Понедельник"
        tsched = [s for s in schedule if s.get("Учитель") == teacher_name and s.get("День") == today_val]
        
        # Format schedule
        formatted_sched = []
        for s in sorted(tsched, key=lambda x: str(x.get("Урок"))):
            lesson_num = s.get("Урок")
            time_map = {1: "08:00", 2: "08:45", 3: "09:40", 4: "10:35", 5: "11:30", 6: "12:25"}
            formatted_sched.append({
                "time": time_map.get(lesson_num, f"Урок {lesson_num}"),
                "subject": s.get("Предмет", "-"),
                "class_name": s.get("Класс", "-"),
                "room": s.get("Кабинет", "-"),
                "is_substitution": False
            })

        return {
            "name": teacher_name,
            "role": teacher_info.get("Должность", "Учитель") if teacher_info else "Учитель",
            "schedule": formatted_sched,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"name": "Ошибка загрузки", "role": "Учитель", "schedule": []}

class SubstitutionRequest(BaseModel):
    teacher_name: str
    date: Optional[str] = None


class AbsenceEventRequest(BaseModel):
    teacher_name: str
    day: Optional[str] = None
    reason: Optional[str] = None
    source: str = "dashboard"
    raw_message: Optional[str] = None

class OrderGenerationRequest(BaseModel):
    missing_teacher: str
    substitute_teacher: str
    lesson_number: int
    class_name: str
    room: str

@router.post("/substitute")
async def schedule_substitute(req: SubstitutionRequest, db: Session = Depends(get_db)):
    plan = find_substitution(req.teacher_name, req.date, db)
    return {"substitution_plan": plan}


@router.post("/absence-event")
async def handle_absence_event(req: AbsenceEventRequest, db: Session = Depends(get_db)):
    result = process_teacher_absence_event(
        teacher_name=req.teacher_name,
        target_day=req.day,
        reason=req.reason,
        source=req.source,
        raw_message=req.raw_message,
        db=db,
    )
    return result

@router.post("/generate-order")
async def generate_order(req: OrderGenerationRequest):
    order = generate_substitution_order(
        req.missing_teacher,
        req.substitute_teacher,
        req.lesson_number,
        req.class_name,
        req.room
    )
    return order

@router.post("/generate-order-pdf")
async def generate_order_pdf(req: OrderGenerationRequest):
    order = generate_substitution_order(
        req.missing_teacher,
        req.substitute_teacher,
        req.lesson_number,
        req.class_name,
        req.room
    )

    current_date = datetime.now().strftime("%d.%m.%Y")

    html_content = create_order_html(
        missing_teacher=req.missing_teacher,
        substitute_teacher=req.substitute_teacher,
        lesson_number=req.lesson_number,
        class_name=req.class_name,
        room=req.room,
        order_date=current_date,
        order_number=f"78-{req.lesson_number}",
        ai_preamble=order.preamble,
        ai_body=order.order_body_ru,
        ai_body_kz=order.order_body_kz,
    )

    return HTMLResponse(content=html_content, status_code=200)


# ─── Heatmap ──────────────────────────────────────────────────────────────────

@router.get("/heatmap")
async def get_heatmap(db: Session = Depends(get_db)):
    """Тепловая карта нагрузки учителей. load=0.0 (свободен) — 1.0 (перегружен)."""
    checker = ConflictChecker(db)
    return checker.get_heatmap_data()


@router.get("/conflicts")
async def get_conflicts(db: Session = Depends(get_db)):
    """Список конфликтов в текущем расписании."""
    checker = ConflictChecker(db)
    conflicts = checker.find_conflicts_in_schedule()
    return {"conflicts": conflicts, "count": len(conflicts)}


class CheckConflictRequest(BaseModel):
    entry_id: int
    new_time_slot_id: int
    new_room_id: Optional[int] = None


def collect_schedule_conflicts(
    entry: ScheduleEntry,
    new_time_slot_id: int,
    db: Session,
    new_room_id: Optional[int] = None,
) -> list:
    checker = ConflictChecker(db)
    found_conflicts = []
    slot = db.query(TimeSlot).get(new_time_slot_id)

    if entry.is_lenta:
        found_conflicts.append(
            {
                "type": "locked_lenta",
                "description": "Ленточные блоки нельзя переносить вручную. Используйте управление лентами.",
            }
        )
        return found_conflicts

    if entry.teacher_id and not checker.check_teacher_availability(
        entry.teacher_id, new_time_slot_id, exclude_entry_id=entry.id
    ):
        teacher = db.query(Teacher).get(entry.teacher_id)
        found_conflicts.append(
            {
                "type": "teacher_busy",
                "description": f"Учитель {teacher.short_name if teacher else '?'} уже занят в {slot.day if slot else '?'} урок {slot.lesson_number if slot else '?'}",
            }
        )

    room_id = new_room_id or entry.room_id
    if room_id and not checker.check_room_availability(
        room_id, new_time_slot_id, exclude_entry_id=entry.id
    ):
        room = db.query(Room).get(room_id)
        found_conflicts.append(
            {
                "type": "room_busy",
                "description": f"Кабинет {room.number if room else '?'} занят в {slot.day if slot else '?'} урок {slot.lesson_number if slot else '?'}",
            }
        )

    if entry.class_id:
        class_conflict = (
            db.query(ScheduleEntry)
            .filter(
                ScheduleEntry.class_id == entry.class_id,
                ScheduleEntry.time_slot_id == new_time_slot_id,
                ScheduleEntry.id != entry.id,
            )
            .first()
        )
        if class_conflict:
            class_name = entry.class_.name if entry.class_ else "класс"
            subject = class_conflict.subject or "другое занятие"
            found_conflicts.append(
                {
                    "type": "class_busy",
                    "description": f"Класс {class_name} уже занят в этом слоте: {subject}",
                }
            )

        parallel_id = entry.class_.parallel_id if entry.class_ else None
        if parallel_id:
            parallel_lenta = (
                db.query(ScheduleEntry)
                .join(LentaGroup, ScheduleEntry.lenta_group_id == LentaGroup.id)
                .filter(
                    ScheduleEntry.time_slot_id == new_time_slot_id,
                    ScheduleEntry.id != entry.id,
                    ScheduleEntry.is_lenta == True,
                    LentaGroup.parallel_id == parallel_id,
                )
                .first()
            )
            if parallel_lenta:
                found_conflicts.append(
                    {
                        "type": "parallel_lenta_busy",
                        "description": "В этом слоте у параллели уже стоит лента, обычный урок сюда переносить нельзя.",
                    }
                )

    return found_conflicts


@router.post("/check-conflict-v2")
async def check_conflict_v2(req: CheckConflictRequest, db: Session = Depends(get_db)):
    entry = db.query(ScheduleEntry).get(req.entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Запись расписания не найдена")

    found_conflicts = collect_schedule_conflicts(
        entry, req.new_time_slot_id, db, req.new_room_id
    )
    return {"ok": len(found_conflicts) == 0, "conflicts": found_conflicts}


@router.patch("/entry-v2/{entry_id}")
async def move_schedule_entry_v2(
    entry_id: int,
    new_time_slot_id: int,
    new_room_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    entry = db.query(ScheduleEntry).get(entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    conflicts = collect_schedule_conflicts(entry, new_time_slot_id, db, new_room_id)
    if conflicts:
        detail = "; ".join(conflict["description"] for conflict in conflicts)
        raise HTTPException(status_code=409, detail=detail)

    entry.time_slot_id = new_time_slot_id
    if new_room_id:
        entry.room_id = new_room_id
    db.commit()
    return {"ok": True, "entry_id": entry_id}


@router.post("/check-conflict")
async def check_conflict(req: CheckConflictRequest, db: Session = Depends(get_db)):
    """
    Проверяет, можно ли переместить урок в новый слот (для Drag-and-Drop).
    Возвращает {ok: bool, conflicts: list}.
    """
    entry = db.query(ScheduleEntry).get(req.entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Запись расписания не найдена")

    checker = ConflictChecker(db)
    found_conflicts = []

    # Проверяем учителя
    if entry.teacher_id:
        teacher_free = checker.check_teacher_availability(
            entry.teacher_id, req.new_time_slot_id, exclude_entry_id=req.entry_id
        )
        if not teacher_free:
            teacher = db.query(Teacher).get(entry.teacher_id)
            slot = db.query(TimeSlot).get(req.new_time_slot_id)
            found_conflicts.append({
                "type": "teacher_busy",
                "description": f"Учитель {teacher.short_name if teacher else '?'} уже занят в {slot.day if slot else '?'} урок {slot.lesson_number if slot else '?'}",
            })

    # Проверяем кабинет
    room_id = req.new_room_id or entry.room_id
    if room_id:
        room_free = checker.check_room_availability(
            room_id, req.new_time_slot_id, exclude_entry_id=req.entry_id
        )
        if not room_free:
            room = db.query(Room).get(room_id)
            slot = db.query(TimeSlot).get(req.new_time_slot_id)
            found_conflicts.append({
                "type": "room_busy",
                "description": f"Кабинет {room.number if room else '?'} занят в {slot.day if slot else '?'} урок {slot.lesson_number if slot else '?'}",
            })

    return {"ok": len(found_conflicts) == 0, "conflicts": found_conflicts}


@router.patch("/entry/{entry_id}")
async def move_schedule_entry(
    entry_id: int,
    new_time_slot_id: int,
    new_room_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Перемещает урок в новый слот (применяется после успешного check-conflict)."""
    entry = db.query(ScheduleEntry).get(entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    # Финальная проверка конфликтов перед сохранением
    checker = ConflictChecker(db)
    conflicts = []
    if entry.teacher_id:
        if not checker.check_teacher_availability(entry.teacher_id, new_time_slot_id, entry_id):
            conflicts.append("teacher_busy")
    room_id = new_room_id or entry.room_id
    if room_id:
        if not checker.check_room_availability(room_id, new_time_slot_id, entry_id):
            conflicts.append("room_busy")
    if conflicts:
        raise HTTPException(status_code=409, detail=f"Конфликт: {', '.join(conflicts)}")

    entry.time_slot_id = new_time_slot_id
    if new_room_id:
        entry.room_id = new_room_id
    db.commit()
    return {"ok": True, "entry_id": entry_id}


# ─── Ленты ────────────────────────────────────────────────────────────────────

class LentaCreateRequest(BaseModel):
    parallel_id: int
    time_slot_id: int


def _validate_lenta_groups(lenta_groups: List[LentaGroup]) -> List[str]:
    conflicts = []
    teacher_ids = set()
    room_ids = set()

    for group in lenta_groups:
        if not group.teacher_id:
            conflicts.append(f"Для группы {group.group_name} не назначен учитель")
        elif group.teacher_id in teacher_ids:
            conflicts.append(f"Учитель назначен сразу в несколько групп ({group.group_name})")
        else:
            teacher_ids.add(group.teacher_id)

        if not group.room_id:
            conflicts.append(f"Для группы {group.group_name} не назначен кабинет")
        elif group.room_id in room_ids:
            conflicts.append(f"Кабинет назначен сразу в несколько групп ({group.group_name})")
        else:
            room_ids.add(group.room_id)

    return conflicts


def _get_parallel_class_ids(parallel: Parallel) -> List[int]:
    return [cls.id for cls in parallel.classes]


@router.get("/parallels")
async def get_parallels(db: Session = Depends(get_db)):
    """Список всех параллелей с их лентами."""
    parallels = db.query(Parallel).order_by(Parallel.grade).all()
    result = []
    for p in parallels:
        groups = db.query(LentaGroup).filter_by(parallel_id=p.id).all()
        result.append({
            "id": p.id,
            "grade": p.grade,
            "has_lenta": p.has_lenta,
            "lenta_subject": p.lenta_subject,
            "classes": [c.name for c in p.classes],
            "lenta_groups": [
                {
                    "id": g.id,
                    "group_name": g.group_name,
                    "level": g.level,
                    "teacher": g.teacher.short_name if g.teacher else None,
                    "room": g.room.number if g.room else None,
                }
                for g in groups
            ],
        })
    return {"parallels": result}


@router.get("/time-slots")
async def get_time_slots(db: Session = Depends(get_db)):
    """Список временных слотов с реальными ID из БД."""
    slots = db.query(TimeSlot).order_by(TimeSlot.id).all()
    return {
        "time_slots": [
            {
                "id": slot.id,
                "day": slot.day,
                "lesson": slot.lesson_number,
                "start_time": slot.start_time,
                "end_time": slot.end_time,
            }
            for slot in slots
        ]
    }


@router.post("/lenta-v2")
async def create_lenta_slot_v2(req: LentaCreateRequest, db: Session = Depends(get_db)):
    parallel = db.query(Parallel).get(req.parallel_id)
    if not parallel:
        raise HTTPException(status_code=404, detail="Параллель не найдена")
    if not parallel.has_lenta:
        raise HTTPException(status_code=400, detail="У этой параллели нет системы лент")

    lenta_groups = db.query(LentaGroup).filter_by(parallel_id=req.parallel_id).all()
    if not lenta_groups:
        raise HTTPException(status_code=400, detail="Нет групп лент для этой параллели")

    slot = db.query(TimeSlot).get(req.time_slot_id)
    if not slot:
        raise HTTPException(status_code=404, detail="Временной слот не найден")

    checker = ConflictChecker(db)
    conflicts = _validate_lenta_groups(lenta_groups)
    class_ids = _get_parallel_class_ids(parallel)

    existing_lenta = (
        db.query(ScheduleEntry)
        .filter(
            ScheduleEntry.is_lenta == True,
            ScheduleEntry.time_slot_id == req.time_slot_id,
            ScheduleEntry.lenta_group_id.isnot(None),
        )
        .join(LentaGroup, ScheduleEntry.lenta_group_id == LentaGroup.id)
        .filter(LentaGroup.parallel_id == req.parallel_id)
        .first()
    )
    if existing_lenta:
        raise HTTPException(status_code=409, detail="Лента уже создана в этом слоте")

    parallel_conflict = (
        db.query(ScheduleEntry)
        .filter(
            ScheduleEntry.time_slot_id == req.time_slot_id,
            ScheduleEntry.class_id.in_(class_ids),
        )
        .first()
    )
    if parallel_conflict:
        conflicts.append("Один из классов параллели уже занят в этом слоте")

    for group in lenta_groups:
        if group.teacher_id and not checker.check_teacher_availability(group.teacher_id, req.time_slot_id):
            teacher = db.query(Teacher).get(group.teacher_id)
            conflicts.append(f"Учитель {teacher.short_name if teacher else '?'} занят ({group.group_name})")
        if group.room_id and not checker.check_room_availability(group.room_id, req.time_slot_id):
            room = db.query(Room).get(group.room_id)
            conflicts.append(f"Кабинет {room.number if room else '?'} занят ({group.group_name})")

    if conflicts:
        return {
            "ok": False,
            "conflicts": conflicts,
            "message": "Обнаружены конфликты — лента не создана",
        }

    created_entries = {"class_blocks": [], "groups": []}

    for cls in parallel.classes:
        db.add(
            ScheduleEntry(
                class_id=cls.id,
                teacher_id=None,
                room_id=None,
                time_slot_id=req.time_slot_id,
                subject=parallel.lenta_subject or "Английский язык",
                is_lenta=True,
                lenta_group_id=None,
            )
        )
        created_entries["class_blocks"].append({"class": cls.name})

    for group in lenta_groups:
        db.add(
            ScheduleEntry(
                class_id=None,
                teacher_id=group.teacher_id,
                room_id=group.room_id,
                time_slot_id=req.time_slot_id,
                subject=parallel.lenta_subject or "Английский язык",
                is_lenta=True,
                lenta_group_id=group.id,
            )
        )
        created_entries["groups"].append(
            {
                "group": group.group_name,
                "teacher": group.teacher.short_name if group.teacher else None,
                "room": group.room.number if group.room else None,
            }
        )

    db.commit()

    return {
        "ok": True,
        "message": f"Лента создана: {slot.day} урок {slot.lesson_number} — {len(lenta_groups)} групп, {len(parallel.classes)} классов заблокированы",
        "slot": f"{slot.day} {slot.start_time}",
        "groups_created": len(lenta_groups),
        "class_blocks_created": len(parallel.classes),
        "entries": created_entries,
    }


@router.delete("/lenta-v2/{parallel_id}/{time_slot_id}")
async def delete_lenta_slot_v2(
    parallel_id: int, time_slot_id: int, db: Session = Depends(get_db)
):
    class_entries = (
        db.query(ScheduleEntry)
        .join(Class, ScheduleEntry.class_id == Class.id)
        .filter(
            ScheduleEntry.is_lenta == True,
            ScheduleEntry.time_slot_id == time_slot_id,
            Class.parallel_id == parallel_id,
        )
        .all()
    )

    group_entries = (
        db.query(ScheduleEntry)
        .join(LentaGroup, ScheduleEntry.lenta_group_id == LentaGroup.id)
        .filter(
            ScheduleEntry.is_lenta == True,
            ScheduleEntry.time_slot_id == time_slot_id,
            LentaGroup.parallel_id == parallel_id,
        )
        .all()
    )

    deleted_ids = set()
    for entry in class_entries + group_entries:
        if entry.id not in deleted_ids:
            deleted_ids.add(entry.id)
            db.delete(entry)

    db.commit()
    return {"ok": True, "deleted_entries": len(deleted_ids)}


@router.post("/lenta")
async def create_lenta_slot(req: LentaCreateRequest, db: Session = Depends(get_db)):
    """
    Создаёт ленту: бронирует один временной слот для всей параллели.
    Проверяет конфликты учителей и кабинетов перед созданием.
    """
    parallel = db.query(Parallel).get(req.parallel_id)
    if not parallel:
        raise HTTPException(status_code=404, detail="Параллель не найдена")
    if not parallel.has_lenta:
        raise HTTPException(status_code=400, detail="У этой параллели нет системы лент")

    lenta_groups = db.query(LentaGroup).filter_by(parallel_id=req.parallel_id).all()
    if not lenta_groups:
        raise HTTPException(status_code=400, detail="Нет групп лент для этой параллели")

    slot = db.query(TimeSlot).get(req.time_slot_id)
    if not slot:
        raise HTTPException(status_code=404, detail="Временной слот не найден")

    checker = ConflictChecker(db)
    conflicts = []

    # Проверяем конфликты для каждой группы
    for group in lenta_groups:
        if group.teacher_id:
            if not checker.check_teacher_availability(group.teacher_id, req.time_slot_id):
                teacher = db.query(Teacher).get(group.teacher_id)
                conflicts.append(f"Учитель {teacher.short_name if teacher else '?'} занят ({group.group_name})")
        if group.room_id:
            if not checker.check_room_availability(group.room_id, req.time_slot_id):
                room = db.query(Room).get(group.room_id)
                conflicts.append(f"Кабинет {room.number if room else '?'} занят ({group.group_name})")

    if conflicts:
        return {
            "ok": False,
            "conflicts": conflicts,
            "message": "Обнаружены конфликты — лента не создана",
        }

    # Создаём записи расписания для каждого класса параллели
    created_entries = []
    for cls in parallel.classes:
        for group in lenta_groups:
            entry = ScheduleEntry(
                class_id=cls.id,
                teacher_id=group.teacher_id,
                room_id=group.room_id,
                time_slot_id=req.time_slot_id,
                subject=parallel.lenta_subject or "Английский язык",
                is_lenta=True,
                lenta_group_id=group.id,
            )
            db.add(entry)
            db.flush()
            created_entries.append({
                "class": cls.name,
                "group": group.group_name,
                "teacher": group.teacher.short_name if group.teacher else None,
                "room": group.room.number if group.room else None,
            })

    db.commit()

    return {
        "ok": True,
        "message": f"Лента создана: {slot.day} урок {slot.lesson_number} — {len(lenta_groups)} групп, {len(parallel.classes)} классов заблокированы",
        "slot": f"{slot.day} {slot.start_time}",
        "groups_created": len(lenta_groups),
        "entries": created_entries,
    }


@router.delete("/lenta/{parallel_id}/{time_slot_id}")
async def delete_lenta_slot(
    parallel_id: int, time_slot_id: int, db: Session = Depends(get_db)
):
    """Удаляет ленту для параллели из указанного слота."""
    deleted = (
        db.query(ScheduleEntry)
        .filter(
            ScheduleEntry.is_lenta == True,
            ScheduleEntry.time_slot_id == time_slot_id,
        )
        .join(Class, ScheduleEntry.class_id == Class.id)
        .filter(Class.parallel_id == parallel_id)
        .all()
    )
    count = len(deleted)
    for entry in deleted:
        db.delete(entry)
    db.commit()
    return {"ok": True, "deleted_entries": count}


# ─── Staff Schedule ───────────────────────────────────────────────────────────

@router.get("/staff")
async def list_staff(db: Session = Depends(get_db)):
    """Список всех сотрудников для выпадающего списка."""
    return {"staff": get_all_staff_list(db)}


@router.get("/staff/{staff_id}/week")
async def get_staff_schedule(staff_id: int, db: Session = Depends(get_db)):
    """Недельное расписание конкретного сотрудника."""
    return get_staff_weekly_schedule(staff_id, db)


class InjectTaskRequest(BaseModel):
    staff_id: int
    task_reminder_id: int


@router.post("/staff/inject-task")
async def inject_task(req: InjectTaskRequest, db: Session = Depends(get_db)):
    """Вставляет задачу в ближайший свободный слот сотрудника."""
    result = inject_task_into_schedule(req.staff_id, req.task_reminder_id, db)
    if not result:
        raise HTTPException(status_code=409, detail="Нет свободных слотов у сотрудника")
    return {"ok": True, "scheduled": result}


# ─── DB-based schedule view ───────────────────────────────────────────────────

@router.get("/db-schedule")
async def get_db_schedule(
    class_name: Optional[str] = None,
    day: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Возвращает расписание из БД (в отличие от generate-schedule который генерирует на лету).
    Можно фильтровать по классу и дню.
    """
    q = db.query(ScheduleEntry)
    if class_name:
        cls = db.query(Class).filter_by(name=class_name).first()
        if cls:
            q = q.filter(ScheduleEntry.class_id == cls.id)
    if day:
        slots = db.query(TimeSlot).filter(TimeSlot.day == day).all()
        slot_ids = [s.id for s in slots]
        q = q.filter(ScheduleEntry.time_slot_id.in_(slot_ids))

    entries = q.all()
    result = []
    for e in entries:
        result.append({
            "id": e.id,
            "time_slot_id": e.time_slot_id,
            "class": e.class_.name if e.class_ else None,
            "teacher": e.teacher.short_name if e.teacher else None,
            "room": e.room.number if e.room else None,
            "day": e.time_slot.day if e.time_slot else None,
            "lesson": e.time_slot.lesson_number if e.time_slot else None,
            "start_time": e.time_slot.start_time if e.time_slot else None,
            "subject": e.subject,
            "is_lenta": e.is_lenta,
            "lenta_group": e.lenta_group.group_name if e.lenta_group else None,
            "is_substitution": e.is_substitution,
        })
    return {"schedule": result, "total": len(result)}

